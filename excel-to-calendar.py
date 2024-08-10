import re
from flask import Flask, render_template, request, send_file
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment
import io

app = Flask(__name__)

new_schedule = Workbook()
schedule_sheet = new_schedule.active
days_of_week = ["Mon", "Tue", "Wed", "Thu", "Fri"]
times = [
    "8:00 a.m.", "8:30 a.m.", "9:00 a.m.", "9:30 a.m.", "10:00 a.m.",
    "10:30 a.m.", "11:00 a.m.", "11:30 a.m.", "12:00 p.m.", "12:30 p.m.",
    "1:00 p.m.", "1:30 p.m.", "2:00 p.m.", "2:30 p.m.", "3:00 p.m.",
    "3:30 p.m.", "4:00 p.m.", "4:30 p.m.", "5:00 p.m.", "5:30 p.m.",
    "6:00 p.m.", "6:30 p.m.", "7:00 p.m.", "7:30 p.m.", "8:00 p.m.",
    "8:30 p.m.", "9:00 p.m.", "9:30 p.m.", "10:00 p.m."
]
term1 = {}
term2 = {}


def load_schedule(file_name):
    wb = xl.load_workbook(file_name)
    sheet = wb["View My Courses"]
    return sheet


# adjusts each cell's height and width
def create_calendar_measurements():
    create_headings()
    for row in range(len(times) + 1):
        schedule_sheet.row_dimensions[row].height = 30

    for col in range(1, 7):
        col_letter = chr(64 + col)
        schedule_sheet.column_dimensions[col_letter].width = 25

    for col in range(1, 7):
        for row in range(1, len(times) + 2):
            cell = schedule_sheet.cell(row, col)
            cell.alignment = Alignment(wrap_text=True)


# creates day headings for calendar
def create_headings():
    schedule_sheet.cell(1, 2).value = "Monday"
    schedule_sheet.cell(1, 3).value = "Tuesday"
    schedule_sheet.cell(1, 4).value = "Wednesday"
    schedule_sheet.cell(1, 5).value = "Thursday"
    schedule_sheet.cell(1, 6).value = "Friday"

    for row_num, time in enumerate(times, start=2):
        cell = schedule_sheet.cell(row=row_num, column=1)
        cell.value = time


def align_top():
    for row_num in range(1, len(times) + 2):
        schedule_sheet.cell(row_num, 1).alignment = Alignment("general", "top")


# examines meeting day cell on excel sheets and extracts the meeting days as a list
def find_class_days(meeting_day_value):
    if meeting_day_value is None:
        return None
    else:
        days = []
        for day in range(5):
            if days_of_week[day] in meeting_day_value:
                days.append(days_of_week[day])
    return days


# returns corresponding column number based on day
def return_day_column(day):
    if day == "Mon":
        return 2
    elif day == "Tue":
        return 3
    elif day == "Wed":
        return 4
    elif day == "Thu":
        return 5
    elif day == "Fri":
        return 6


# places each course in the calendar based on its days and time
def place_course_in_calendar(class_name, days, sheet):
    meeting_day_value = get_meeting_day_value(class_name, sheet)
    class_times = get_class_times(meeting_day_value)
    if days and class_times is not None:
        start_time = class_times[0]
        end_time = class_times[1]
        start_row = times.index(start_time) + 2
        end_row = times.index(end_time) + 2
        title = f'''{class_name}
{start_time} - {end_time}'''
        for day in days:
            column = return_day_column(day)
            schedule_sheet.merge_cells(
                start_row=start_row, start_column=column, end_row=end_row - 1, end_column=column)
            cell = schedule_sheet.cell(
                row=start_row, column=column, value=title)
            cell.alignment = Alignment(vertical='center', horizontal='center')


# returns meeting day value of the class
def get_meeting_day_value(class_name, sheet):
    for row in range(4, sheet.max_row + 1):
        if class_name == sheet.cell(row, 5).value:
            return sheet.cell(row, 8).value

    return None


def find_term(meeting_day_value):
    if meeting_day_value is not None:
        if "2024-09" in meeting_day_value:
            return 1
        elif "2025-01" in meeting_day_value:
            return 2


def create_term_dicts(sheet):
    for row in range(4, sheet.max_row + 1):
        section_name = sheet.cell(row, 5).value
        days = find_class_days(sheet.cell(row, 8).value)
        term = find_term(sheet.cell(row, 8).value)
        if term == 1:
            term1[section_name] = days
        elif term == 2:
            term2[section_name] = days


def get_class_times(meeting_day_value):
    if meeting_day_value is not None:
        # Extract the meeting time using a regular expression
        meeting_time_pattern = r"\|\s*(\d{1,2}:\d{2}\s*(?:a\.m\.|p\.m\.)\s*-\s*\d{1,2}:\d{2}\s*(?:a\.m\.|p\.m\.))\s*\|"
        match = re.search(meeting_time_pattern, meeting_day_value)

        if match:
            meeting_time = match.group(1)
            start_time, end_time = meeting_time.split(" - ")
            class_times = [start_time, end_time]
            return class_times
    else:
        return None


def create_schedule(file_name, term_num):
    sheet = load_schedule(file_name)
    create_term_dicts(sheet)

    if term_num == 1:
        term_num = term1
    else:
        term_num = term2

    for school_class in term_num:
        if school_class is not None:
            place_course_in_calendar(
                school_class, term_num.get(school_class), sheet)

    create_calendar_measurements()
    align_top()
    new_schedule.save("new-schedule.xlsx")


@app.route('/')
def index():
    return render_template('ubc-excel-to-schedule.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    global new_schedule, schedule_sheet

    if 'file' not in request.files:
        return 'No file part', 400

    file = request.files['file']
    if file.filename == '':
        return 'No selected file', 400
    
    new_schedule = Workbook()
    schedule_sheet = new_schedule.active
    
    term_num = int(request.form['term_num'])
   
    if not ("View_My_Courses") in file.filename:
        return "Excel file must have View_My_Courses as the name. Press the back arrow to resubmit an excel file."
    create_schedule(file, term_num)

    output = io.BytesIO()
    new_schedule.save(output)
    output.seek(0)

    return send_file(output, download_name='modified_term' + str(term_num) + "schedule.xlsx", as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)
